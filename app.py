from __future__ import annotations

import csv
import hashlib
from http.cookies import SimpleCookie
import io
import json
import logging
from logging.handlers import RotatingFileHandler
import mimetypes
import os
import ipaddress
import re
import shutil
import socket
import sqlite3
import sys
import tempfile
import time
import uuid
from datetime import date, datetime
from email.parser import BytesParser
from email.policy import default
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse
from urllib.request import Request, urlopen

from config import SETTINGS
from auth_service import (
    authenticate,
    cleanup_sessions,
    create_session,
    create_user,
    public_user,
    revoke_session,
    session_user,
)

BASE_DIR = SETTINGS.base_dir
STATIC_DIR = SETTINGS.static_dir
DATA_DIR = SETTINGS.data_dir
LEGACY_ATTACHMENTS_DIR = DATA_DIR / "attachments"
ATTACHMENTS_DIR = SETTINGS.originals_dir
DB_PATH = SETTINGS.db_path
MAX_UPLOAD_BYTES = SETTINGS.max_upload_bytes
LOGGER = logging.getLogger("qoyod_archive")
SESSION_COOKIE = "qoyod_session"
IMPORTANT_AUDIT_ACTIONS = {
    "رفع مرفق",
    "حذف مرفق",
    "فتح القيد",
    "فتح ملف مرفق",
}

STATUSES = [
    "بانتظار المرفقات",
    "تم رفع المرفقات",
    "مؤرشف",
]

WORKFLOW_ACTIONS = {
    "archive": "مؤرشف",
    "unarchive": "تم رفع المرفقات",
}

ACCOUNTANT_DISPLAY_NAMES = {
    "a.mobarak": "ا. أحمد مبارك",
    "أحمد.مبارك": "ا. أحمد مبارك",
    "احمد.مبارك": "ا. أحمد مبارك",
    "m.salah": "ا. محمد صلاح",
    "mo.salah": "ا. محمد صلاح",
    "محمد.صلاح": "ا. محمد صلاح",
    "mohsen": "ا. محسن اشرف",
    "محسن.اشرف": "ا. محسن اشرف",
    "sherif": "ا. شريف عزام",
    "شريف.عزام": "ا. شريف عزام",
    "omar.adel": "ا. عمر عادل",
    "عمر.عادل": "ا. عمر عادل",
    "ali": "ا. علي البنا",
    "علي.البنا": "ا. علي البنا",
    "b.elabd": "ا. بلال العبد",
    "بلال.العبد": "ا. بلال العبد",
    "eslam": "ا. اسلام غنيم",
    "اسلام.غنيم": "ا. اسلام غنيم",
    "hazzem": "ا. حازم عصام",
    "حازم.عصام": "ا. حازم عصام",
    "i.akram": "ا. ابراهيم اكرم",
    "ابراهيم.اكرم": "ا. ابراهيم اكرم",
    "m.ashraf": "ا. محمد اشرف",
    "محمد.اشرف": "ا. محمد اشرف",
    "m.yahya": "ا. محمود يحي",
    "محمود.يحي": "ا. محمود يحي",
    "tolba": "ا. احمد طلبه",
    "احمد.طلبه": "ا. احمد طلبه",
}


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat(sep=" ")


def normalize_accountant(value: object) -> str:
    raw = re.sub(r"\s+", " ", str(value or "").strip())
    raw_without_title = re.sub(r"^[اأإ]\s*\.\s*", "", raw)
    lookup_key = re.sub(
        r"[\s._-]+", ".", raw_without_title.casefold()).strip(".")
    return ACCOUNTANT_DISPLAY_NAMES.get(lookup_key, raw)


def normalize_paper_marker(value: object) -> bool | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    normalized = raw.casefold()
    if raw in {"☑", "☒", "✓", "✔"} or normalized in {
        "1",
        "true",
        "yes",
        "checked",
        "present",
        "موجود",
        "نعم",
    }:
        return True
    if raw in {"☐", "□"} or normalized in {
        "0",
        "false",
        "no",
        "unchecked",
        "missing",
        "غير موجود",
        "لا",
    }:
        return False
    return None


def setup_storage() -> None:
    for path in (
        DATA_DIR,
        LEGACY_ATTACHMENTS_DIR,
        ATTACHMENTS_DIR,
        SETTINGS.scans_dir,
        SETTINGS.excel_dir,
        SETTINGS.processed_dir,
        SETTINGS.reports_dir,
        SETTINGS.logs_dir,
    ):
        path.mkdir(parents=True, exist_ok=True)


def setup_logging() -> None:
    setup_storage()
    if LOGGER.handlers:
        return
    LOGGER.setLevel(getattr(logging, SETTINGS.log_level, logging.INFO))
    formatter = logging.Formatter(
        "%(asctime)s %(levelname)s %(name)s %(message)s")
    file_handler = RotatingFileHandler(
        SETTINGS.logs_dir / "archive.log",
        maxBytes=5 * 1024 * 1024,
        backupCount=5,
        encoding="utf-8",
    )
    file_handler.setFormatter(formatter)
    LOGGER.addHandler(file_handler)
    console = logging.StreamHandler()
    console.setFormatter(formatter)
    LOGGER.addHandler(console)


def connect_db() -> sqlite3.Connection:
    setup_storage()
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA busy_timeout = 30000")
    return conn


def ensure_column(conn: sqlite3.Connection, table: str, column: str, definition: str) -> None:
    columns = {row["name"]
               for row in conn.execute(f"PRAGMA table_info({table})")}
    if column not in columns:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def init_db() -> None:
    with connect_db() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS entries (
                trans_no TEXT PRIMARY KEY,
                entry_type TEXT DEFAULT '',
                entry_date TEXT DEFAULT '',
                num TEXT DEFAULT '',
                name TEXT DEFAULT '',
                amount REAL DEFAULT 0,
                accountant TEXT DEFAULT '',
                status TEXT NOT NULL DEFAULT 'بانتظار المرفقات',
                file_box TEXT DEFAULT '',
                shelf TEXT DEFAULT '',
                paper_range TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        for column, definition in (
            ("project", "TEXT DEFAULT ''"),
            ("cost_center", "TEXT DEFAULT ''"),
            ("supplier", "TEXT DEFAULT ''"),
            ("client", "TEXT DEFAULT ''"),
            ("description", "TEXT DEFAULT ''"),
            ("match_status", "TEXT DEFAULT ''"),
            ("risk_score", "INTEGER DEFAULT 0"),
            ("reviewer_status", "TEXT DEFAULT ''"),
            ("archive_date", "TEXT DEFAULT ''"),
            ("paper_received", "INTEGER DEFAULT 0"),
            ("paper_received_at", "TEXT DEFAULT ''"),
            ("paper_received_by", "TEXT DEFAULT ''"),
            ("paper_source", "TEXT DEFAULT ''"),
        ):
            ensure_column(conn, "entries", column, definition)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trans_no TEXT NOT NULL,
                original_name TEXT NOT NULL,
                stored_name TEXT NOT NULL,
                mime TEXT DEFAULT 'application/octet-stream',
                size INTEGER NOT NULL DEFAULT 0,
                source TEXT NOT NULL DEFAULT 'upload',
                source_url TEXT DEFAULT '',
                uploaded_by TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                FOREIGN KEY (trans_no) REFERENCES entries(trans_no) ON DELETE CASCADE
            )
            """
        )
        for column, definition in (
            ("sha256", "TEXT DEFAULT ''"),
            ("duplicate_of", "INTEGER"),
            ("ocr_text", "TEXT DEFAULT ''"),
            ("ocr_provider", "TEXT DEFAULT ''"),
            ("extraction_status", "TEXT DEFAULT 'لم يبدأ'"),
            ("extraction_json", "TEXT DEFAULT ''"),
        ):
            ensure_column(conn, "attachments", column, definition)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS notes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                trans_no TEXT NOT NULL,
                author TEXT DEFAULT '',
                role TEXT DEFAULT '',
                note TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (trans_no) REFERENCES entries(trans_no) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS excel_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_name TEXT NOT NULL,
                stored_name TEXT DEFAULT '',
                imported_by TEXT DEFAULT '',
                imported_rows INTEGER DEFAULT 0,
                updated_rows INTEGER DEFAULT 0,
                mapping_json TEXT DEFAULT '',
                imported_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS paper_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_name TEXT NOT NULL,
                stored_name TEXT DEFAULT '',
                imported_by TEXT DEFAULT '',
                total_rows INTEGER DEFAULT 0,
                matched_rows INTEGER DEFAULT 0,
                unmatched_rows INTEGER DEFAULT 0,
                imported_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS extracted_fields (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                document_id INTEGER NOT NULL,
                field_name TEXT NOT NULL,
                field_value TEXT DEFAULT '',
                confidence REAL DEFAULT 0,
                is_uncertain INTEGER DEFAULT 0,
                FOREIGN KEY (document_id) REFERENCES attachments(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS matching_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                document_id INTEGER NOT NULL,
                trans_no TEXT NOT NULL,
                match_status TEXT NOT NULL,
                confidence_score INTEGER DEFAULT 0,
                risk_score INTEGER DEFAULT 0,
                reasons_json TEXT DEFAULT '[]',
                reviewed_by TEXT DEFAULT '',
                reviewed_at TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                FOREIGN KEY (document_id) REFERENCES attachments(id) ON DELETE CASCADE,
                FOREIGN KEY (trans_no) REFERENCES entries(trans_no) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_name TEXT DEFAULT '',
                action TEXT NOT NULL,
                entity_type TEXT NOT NULL,
                entity_id TEXT NOT NULL,
                old_value TEXT DEFAULT '',
                new_value TEXT DEFAULT '',
                created_at TEXT NOT NULL
            )
            """
        )
        ensure_column(conn, "audit_logs", "ip_address", "TEXT DEFAULT ''")
        ensure_column(conn, "audit_logs", "user_login", "TEXT DEFAULT ''")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                display_name TEXT NOT NULL,
                password_salt TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user',
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS sessions (
                token_hash TEXT PRIMARY KEY,
                user_id INTEGER NOT NULL,
                ip_address TEXT DEFAULT '',
                created_at TEXT NOT NULL,
                expires_at TEXT NOT NULL,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_entries_status ON entries(status)")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_entries_name ON entries(name)")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_entries_accountant ON entries(accountant)")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_attachments_trans ON attachments(trans_no)")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_attachments_hash ON attachments(sha256)")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_match_trans ON matching_results(trans_no)")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_audit_entity ON audit_logs(entity_type, entity_id)")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_audit_user ON audit_logs(user_name)")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_audit_created ON audit_logs(created_at)")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_users_username ON users(username)")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_sessions_user ON sessions(user_id)")
        conn.execute(
            "DELETE FROM audit_logs WHERE action NOT IN (?, ?, ?, ?)",
            tuple(sorted(IMPORTANT_AUDIT_ACTIONS)),
        )
        cleanup_sessions(conn)
        conn.commit()


def backup_database(destination: Path) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(DB_PATH) as source_conn, sqlite3.connect(destination) as backup_conn:
        source_conn.backup(backup_conn)
        integrity = backup_conn.execute("PRAGMA integrity_check").fetchone()[0]
        if integrity != "ok":
            raise RuntimeError(f"فشل التحقق من النسخة الاحتياطية: {integrity}")
        backup_conn.execute("PRAGMA journal_mode = DELETE").fetchone()
    return destination


def create_daily_backup() -> Path:
    backup_path = DATA_DIR / "backups" / "daily" / \
        f"archive-{datetime.now():%Y-%m-%d}.db"
    if backup_path.exists() and backup_path.stat().st_size > 0:
        return backup_path
    backup_database(backup_path)
    with connect_db() as conn:
        audit(
            conn,
            "إنشاء نسخة احتياطية يومية",
            "backup",
            backup_path.name,
            "النظام",
            "",
            {"path": str(backup_path), "size": backup_path.stat().st_size},
            "127.0.0.1",
        )
        conn.commit()
    return backup_path


def audit(
    conn: sqlite3.Connection,
    action: str,
    entity_type: str,
    entity_id: object,
    user_name: str = "",
    old_value: object = "",
    new_value: object = "",
    ip_address: str = "",
    user_login: str = "",
) -> None:
    if action not in IMPORTANT_AUDIT_ACTIONS:
        return

    def serialize(value: object) -> str:
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False, default=str)
        return str(value or "")

    conn.execute(
        """
        INSERT INTO audit_logs (
            user_name, user_login, action, entity_type, entity_id,
            old_value, new_value, created_at, ip_address
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            user_name or "غير محدد",
            user_login,
            action,
            entity_type,
            str(entity_id),
            serialize(old_value),
            serialize(new_value),
            now_iso(),
            ip_address,
        ),
    )
    LOGGER.info(
        "AUDIT user=%s login=%s ip=%s action=%s entity=%s:%s details=%s",
        user_name or "غير محدد",
        user_login or "-",
        ip_address or "-",
        action,
        entity_type,
        entity_id,
        serialize(new_value)[:500],
    )


def row_to_entry(row: sqlite3.Row) -> dict:
    result = {
        "trans_no": row["trans_no"],
        "entry_type": row["entry_type"] or "",
        "entry_date": row["entry_date"] or "",
        "num": row["num"] or "",
        "name": row["name"] or "",
        "amount": row["amount"] or 0,
        "accountant": row["accountant"] or "",
        "status": row["status"] or STATUSES[0],
        "file_box": row["file_box"] or "",
        "shelf": row["shelf"] or "",
        "paper_range": row["paper_range"] or "",
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
        "attachments_count": row["attachments_count"] if "attachments_count" in row.keys() else 0,
    }
    for name in (
        "project",
        "cost_center",
        "supplier",
        "client",
        "description",
        "match_status",
        "risk_score",
        "reviewer_status",
        "archive_date",
        "paper_received",
        "paper_received_at",
        "paper_received_by",
        "paper_source",
    ):
        result[name] = row[name] if name in row.keys(
        ) and row[name] is not None else ""
    return result


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"[\s_:\-]+", "", text)


def normalize_date(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def normalize_amount(value: object) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value)
    text = text.replace("$", "").replace(
        ",", "").replace("ج", "").replace("EGP", "")
    text = re.sub(r"[^\d.\-]", "", text)
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def pick(row: dict[str, object], *names: str) -> object:
    keys = {normalize_header(key): key for key in row.keys()}
    for name in names:
        key = keys.get(normalize_header(name))
        if key is not None:
            return row.get(key)
    return ""


IMPORT_ALIASES = {
    "trans_no": ("Trans#", "Trans", "Transaction", "رقم القيد", "رقم المستند"),
    "entry_type": ("Type", "نوع", "نوع القيد"),
    "entry_date": ("Date", "التاريخ"),
    "num": ("Num", "No", "رقم المستند", "رقم الفاتورة"),
    "name": ("Name", "الاسم", "الجهة"),
    "amount": ("Amount", "المبلغ", "القيمة", "الإجمالي"),
    "debit": ("Debit", "مدين"),
    "credit": ("Credit", "دائن"),
    "project": ("Project", "المشروع", "اسم المشروع"),
    "cost_center": ("Cost Center", "CostCenter", "مركز التكلفة"),
    "accountant": (
        "Accountant",
        "Last modified by",
        "Last Modified By",
        "Entered by",
        "استلام المحاسب",
        "اسم المحاسب",
        "المحاسب",
    ),
    "paper_received": ("checkbox", "Checkbox", "موجود", "موجود ورقيًا", "استلام ورقي"),
    "supplier": ("Supplier", "المورد", "اسم المورد"),
    "client": ("Client", "Customer", "العميل", "اسم العميل"),
    "description": ("Description", "Memo", "البيان", "الوصف"),
    "status": ("Status", "حالة القيد", "الحالة"),
}

QUICKBOOKS_MAPPING_FIELDS = (
    "paper_received",
    "trans_no",
    "entry_type",
    "accountant",
    "entry_date",
    "num",
    "name",
    "description",
    "debit",
    "credit",
    "amount",
)


def clean_header_row(values: object) -> list[str]:
    headers = [str(cell or "").strip() for cell in values]
    while headers and not headers[-1]:
        headers.pop()
    return headers


def import_headers(filename: str, data: bytes) -> list[str]:
    suffix = Path(filename).suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "قراءة Excel تحتاج مكتبة openpyxl. شغّل setup.bat.") from exc
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
            temp.write(data)
            temp_path = temp.name
        try:
            workbook = load_workbook(temp_path, read_only=True, data_only=True)
            return clean_header_row(next(workbook.active.iter_rows(values_only=True)))
        finally:
            try:
                os.unlink(temp_path)
            except OSError:
                pass
    if suffix == ".csv":
        text = data.decode("utf-8-sig")
        return clean_header_row(next(csv.reader(io.StringIO(text))))
    raise RuntimeError("ارفع ملف Excel بصيغة .xlsx أو ملف CSV.")


def suggested_mapping(headers: list[str]) -> dict[str, str]:
    normalized = {normalize_header(header): header for header in headers}
    result = {}
    for field, aliases in IMPORT_ALIASES.items():
        for alias in aliases:
            if normalize_header(alias) in normalized:
                result[field] = normalized[normalize_header(alias)]
                break
    return result


def parse_import_rows(filename: str, data: bytes, mapping: dict[str, str] | None = None) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    raw_records: list[dict] = []
    mapping = mapping or {}

    def value(raw: dict[str, object], field: str) -> object:
        mapped_header = mapping.get(field, "")
        if mapped_header and mapped_header in raw:
            return raw.get(mapped_header)
        return pick(raw, *IMPORT_ALIASES[field])

    def make_record(raw: dict[str, object]) -> dict | None:
        trans_no = str(value(raw, "trans_no") or "").strip()
        header_values = {normalize_header(alias)
                         for alias in IMPORT_ALIASES["trans_no"]}
        if (
            not trans_no
            or normalize_header(trans_no) in header_values
            or trans_no.startswith("=")
            or trans_no.upper() in {"#REF!", "#VALUE!", "#N/A", "#NAME?", "#NUM!", "#NULL!"}
        ):
            return None
        if re.fullmatch(r"\d+\.0", trans_no):
            trans_no = trans_no[:-2]
        if not re.fullmatch(r"\d+", trans_no):
            return None
        debit = abs(normalize_amount(value(raw, "debit")))
        credit = abs(normalize_amount(value(raw, "credit")))
        direct_amount = abs(normalize_amount(value(raw, "amount")))
        account = str(pick(raw, "Account", "الحساب") or "").strip()
        is_total_row = not account and debit > 0 and credit > 0
        return {
            "trans_no": trans_no,
            "entry_type": str(value(raw, "entry_type") or "").strip(),
            "entry_date": normalize_date(value(raw, "entry_date")),
            "num": str(value(raw, "num") or "").strip(),
            "name": str(value(raw, "name") or "").strip(),
            "amount": direct_amount,
            "project": str(value(raw, "project") or "").strip(),
            "cost_center": str(value(raw, "cost_center") or "").strip(),
            "accountant": normalize_accountant(value(raw, "accountant")),
            "supplier": str(value(raw, "supplier") or "").strip(),
            "client": str(value(raw, "client") or "").strip(),
            "description": str(value(raw, "description") or "").strip(),
            "status": str(value(raw, "status") or "").strip(),
            "_paper_received": normalize_paper_marker(value(raw, "paper_received")),
            "_debit": 0.0 if is_total_row else debit,
            "_credit": 0.0 if is_total_row else credit,
        }

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "قراءة Excel تحتاج مكتبة openpyxl. شغل البرنامج من ملف run.bat المرفق.") from exc

        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
            temp.write(data)
            temp_path = temp.name
        try:
            workbook = load_workbook(temp_path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = clean_header_row(next(rows))
            if len(headers) > 1 and not headers[1] and "استلام المحاسب" in headers:
                headers[1] = "Type"
            for values in rows:
                raw = dict(zip(headers, values[: len(headers)]))
                record = make_record(raw)
                if record:
                    raw_records.append(record)
        finally:
            try:
                os.unlink(temp_path)
            except OSError:
                pass
    elif suffix == ".csv":
        text = data.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for raw in reader:
            record = make_record(raw)
            if record:
                raw_records.append(record)
    else:
        raise RuntimeError("ارفع ملف Excel بصيغة .xlsx أو ملف CSV.")

    records_by_trans: dict[str, dict] = {}
    text_fields = (
        "entry_type",
        "entry_date",
        "num",
        "name",
        "project",
        "cost_center",
        "accountant",
        "supplier",
        "client",
        "description",
        "status",
    )
    for raw_record in raw_records:
        trans_no = raw_record["trans_no"]
        record = records_by_trans.setdefault(
            trans_no,
            {
                "trans_no": trans_no,
                **{field: "" for field in text_fields},
                "amount": 0.0,
                "_debit_total": 0.0,
                "_credit_total": 0.0,
                "_direct_amount": 0.0,
                "_paper_values": set(),
            },
        )
        for field in text_fields:
            if not record[field] and raw_record.get(field):
                record[field] = raw_record[field]
        record["_debit_total"] += float(raw_record.get("_debit") or 0)
        record["_credit_total"] += float(raw_record.get("_credit") or 0)
        record["_direct_amount"] = max(
            record["_direct_amount"],
            abs(float(raw_record.get("amount") or 0)),
        )
        if raw_record.get("_paper_received") is not None:
            record["_paper_values"].add(bool(raw_record["_paper_received"]))

    records = []
    for record in records_by_trans.values():
        record["amount"] = round(
            max(
                record.pop("_direct_amount"),
                record.pop("_debit_total"),
                record.pop("_credit_total"),
            ),
            2,
        )
        paper_values = record.pop("_paper_values")
        record["paper_received"] = True if True in paper_values else False if False in paper_values else None
        records.append(record)
    return records


def parse_paper_register_rows(
    filename: str,
    data: bytes,
    mapping: dict[str, str] | None = None,
) -> list[str]:
    suffix = Path(filename).suffix.lower()
    mapping = mapping or {}
    values: list[str] = []

    def add_row(raw: dict[str, object]) -> None:
        mapped_header = mapping.get("trans_no", "")
        value = raw.get(mapped_header) if mapped_header and mapped_header in raw else pick(
            raw,
            *IMPORT_ALIASES["trans_no"],
            "قيد",
            "رقم",
        )
        trans_no = str(value or "").strip()
        header_values = {normalize_header(alias)
                         for alias in IMPORT_ALIASES["trans_no"]}
        if trans_no and normalize_header(trans_no) not in header_values:
            values.append(trans_no)

    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise RuntimeError(
                "قراءة Excel تحتاج مكتبة openpyxl. شغّل setup.bat.") from exc
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as temp:
            temp.write(data)
            temp_path = temp.name
        try:
            workbook = load_workbook(temp_path, read_only=True, data_only=True)
            rows = workbook.active.iter_rows(values_only=True)
            headers = clean_header_row(next(rows))
            for row_values in rows:
                add_row(dict(zip(headers, row_values[: len(headers)])))
            workbook.close()
        finally:
            try:
                os.unlink(temp_path)
            except OSError:
                pass
    elif suffix == ".csv":
        reader = csv.DictReader(io.StringIO(data.decode("utf-8-sig")))
        for raw in reader:
            add_row(raw)
    else:
        raise RuntimeError("ارفع ملف Excel بصيغة .xlsx أو ملف CSV.")
    return list(dict.fromkeys(values))


def upsert_entry(conn: sqlite3.Connection, payload: dict) -> str:
    trans_no = str(payload.get("trans_no") or "").strip()
    if not trans_no:
        raise ValueError("رقم القيد مطلوب.")
    current = now_iso()
    requested_status = str(payload.get("status") or "").strip()
    conn.execute(
        """
        INSERT INTO entries (
            trans_no, entry_type, entry_date, num, name, amount, accountant,
            status, file_box, shelf, paper_range, project, cost_center, supplier,
            client, description, match_status, risk_score, reviewer_status,
            archive_date, created_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(trans_no) DO UPDATE SET
            entry_type = COALESCE(NULLIF(excluded.entry_type, ''), entries.entry_type),
            entry_date = COALESCE(NULLIF(excluded.entry_date, ''), entries.entry_date),
            num = COALESCE(NULLIF(excluded.num, ''), entries.num),
            name = COALESCE(NULLIF(excluded.name, ''), entries.name),
            amount = CASE WHEN excluded.amount != 0 THEN excluded.amount ELSE entries.amount END,
            accountant = COALESCE(NULLIF(excluded.accountant, ''), entries.accountant),
            status = CASE WHEN ? != '' THEN excluded.status ELSE entries.status END,
            file_box = COALESCE(NULLIF(excluded.file_box, ''), entries.file_box),
            shelf = COALESCE(NULLIF(excluded.shelf, ''), entries.shelf),
            paper_range = COALESCE(NULLIF(excluded.paper_range, ''), entries.paper_range),
            project = COALESCE(NULLIF(excluded.project, ''), entries.project),
            cost_center = COALESCE(NULLIF(excluded.cost_center, ''), entries.cost_center),
            supplier = COALESCE(NULLIF(excluded.supplier, ''), entries.supplier),
            client = COALESCE(NULLIF(excluded.client, ''), entries.client),
            description = COALESCE(NULLIF(excluded.description, ''), entries.description),
            match_status = COALESCE(NULLIF(excluded.match_status, ''), entries.match_status),
            risk_score = CASE WHEN excluded.risk_score != 0 THEN excluded.risk_score ELSE entries.risk_score END,
            reviewer_status = COALESCE(NULLIF(excluded.reviewer_status, ''), entries.reviewer_status),
            archive_date = COALESCE(NULLIF(excluded.archive_date, ''), entries.archive_date),
            updated_at = excluded.updated_at
        """,
        (
            trans_no,
            str(payload.get("entry_type") or "").strip(),
            str(payload.get("entry_date") or "").strip(),
            str(payload.get("num") or "").strip(),
            str(payload.get("name") or "").strip(),
            normalize_amount(payload.get("amount")),
            str(payload.get("accountant") or "").strip(),
            requested_status or STATUSES[0],
            str(payload.get("file_box") or "").strip(),
            str(payload.get("shelf") or "").strip(),
            str(payload.get("paper_range") or "").strip(),
            str(payload.get("project") or "").strip(),
            str(payload.get("cost_center") or "").strip(),
            str(payload.get("supplier") or "").strip(),
            str(payload.get("client") or "").strip(),
            str(payload.get("description") or "").strip(),
            str(payload.get("match_status") or "").strip(),
            int(payload.get("risk_score") or 0),
            str(payload.get("reviewer_status") or "").strip(),
            str(payload.get("archive_date") or "").strip(),
            current,
            current,
            requested_status,
        ),
    )
    return trans_no


def ensure_entry(conn: sqlite3.Connection, trans_no: str) -> None:
    upsert_entry(conn, {"trans_no": trans_no})


def extension_for(filename: str, mime: str) -> str:
    ext = Path(filename).suffix.lower()
    if ext:
        return ext[:12]
    guessed = mimetypes.guess_extension(mime or "")
    return guessed or ".bin"


def attachment_path(row: sqlite3.Row | dict) -> Path:
    stored_name = row["stored_name"]
    for directory in (ATTACHMENTS_DIR, LEGACY_ATTACHMENTS_DIR):
        candidate = (directory / stored_name).resolve()
        if directory.resolve() in candidate.parents and candidate.exists():
            return candidate
    return (ATTACHMENTS_DIR / stored_name).resolve()


def validate_remote_url(url: str) -> None:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.hostname:
        raise ValueError("اكتب رابط HTTP أو HTTPS صحيح.")
    try:
        addresses = socket.getaddrinfo(parsed.hostname, parsed.port or (
            443 if parsed.scheme == "https" else 80))
    except socket.gaierror as exc:
        raise ValueError("تعذر الوصول إلى اسم الموقع.") from exc
    for address in addresses:
        ip = ipaddress.ip_address(address[4][0])
        if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved or ip.is_multicast:
            raise ValueError(
                "لا يمكن تحميل ملفات من عنوان شبكة داخلي أو محجوز.")


def store_attachment(
    conn: sqlite3.Connection,
    trans_no: str,
    original_name: str,
    data: bytes,
    mime: str = "",
    source: str = "upload",
    source_url: str = "",
    uploaded_by: str = "",
    ip_address: str = "",
) -> int:
    if not trans_no:
        raise ValueError("رقم القيد مطلوب.")
    if not data:
        raise ValueError("الملف فارغ.")
    if MAX_UPLOAD_BYTES and len(data) > MAX_UPLOAD_BYTES:
        raise ValueError(
            f"حجم الملف أكبر من الحد المسموح {MAX_UPLOAD_BYTES // 1024 // 1024}MB.")

    if not conn.execute("SELECT 1 FROM entries WHERE trans_no = ?", (trans_no,)).fetchone():
        raise ValueError("رقم القيد غير موجود في كشف الأرشيف المعتمد.")
    mime = mime or mimetypes.guess_type(
        original_name)[0] or "application/octet-stream"
    if not (
        mime == "application/pdf"
        or mime.startswith("image/")
        or Path(original_name).suffix.lower() in {".pdf", ".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff", ".heic"}
    ):
        raise ValueError("المسموح ملفات PDF أو الصور فقط.")
    file_hash = hashlib.sha256(data).hexdigest()
    duplicate = conn.execute(
        "SELECT id FROM attachments WHERE sha256 = ? ORDER BY id LIMIT 1",
        (file_hash,),
    ).fetchone()
    stored_name = f"{trans_no}_{uuid.uuid4().hex}{extension_for(original_name, mime)}"
    path = ATTACHMENTS_DIR / stored_name
    path.write_bytes(data)
    current = now_iso()
    cursor = conn.execute(
        """
        INSERT INTO attachments (
            trans_no, original_name, stored_name, mime, size, source, source_url, uploaded_by,
            created_at, sha256, duplicate_of, extraction_status
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            trans_no,
            original_name or stored_name,
            stored_name,
            mime,
            len(data),
            source,
            source_url,
            uploaded_by,
            current,
            file_hash,
            int(duplicate["id"]) if duplicate else None,
            "بانتظار الاستخراج",
        ),
    )
    conn.execute(
        "UPDATE entries SET status = CASE WHEN status = ? THEN ? ELSE status END, updated_at = ? WHERE trans_no = ?",
        (STATUSES[0], STATUSES[1], current, trans_no),
    )
    attachment_id = int(cursor.lastrowid)
    audit(
        conn,
        "رفع مرفق",
        "attachment",
        attachment_id,
        uploaded_by,
        "",
        {
            "trans_no": trans_no,
            "file_name": original_name,
            "size": len(data),
            "duplicate_of": int(duplicate["id"]) if duplicate else None,
        },
        ip_address,
    )
    return attachment_id


def read_json(handler: BaseHTTPRequestHandler) -> dict:
    length = int(handler.headers.get("Content-Length", "0") or "0")
    if length == 0:
        return {}
    return json.loads(handler.rfile.read(length).decode("utf-8"))


def parse_multipart(handler: BaseHTTPRequestHandler) -> tuple[dict[str, list[str]], list[dict]]:
    content_type = handler.headers.get("Content-Type", "")
    length = int(handler.headers.get("Content-Length", "0") or "0")
    if "multipart/form-data" not in content_type:
        raise ValueError("الطلب ليس multipart/form-data.")
    if MAX_UPLOAD_BYTES and length > MAX_UPLOAD_BYTES * 3:
        raise ValueError("حجم الطلب كبير جدًا.")

    body = handler.rfile.read(length)
    message = BytesParser(policy=default).parsebytes(
        b"Content-Type: " +
        content_type.encode("utf-8") + b"\r\nMIME-Version: 1.0\r\n\r\n" + body
    )
    fields: dict[str, list[str]] = {}
    files: list[dict] = []
    for part in message.iter_parts():
        disposition = part.get("Content-Disposition", "")
        if "form-data" not in disposition:
            continue
        name = part.get_param("name", header="content-disposition") or ""
        filename = part.get_filename()
        payload = part.get_payload(decode=True) or b""
        if filename:
            files.append(
                {
                    "field": name,
                    "filename": filename,
                    "content": payload,
                    "mime": part.get_content_type(),
                }
            )
        else:
            fields.setdefault(name, []).append(payload.decode(
                part.get_content_charset() or "utf-8", errors="replace"))
    return fields, files


def json_bytes(payload: object) -> bytes:
    return json.dumps(payload, ensure_ascii=False).encode("utf-8")


def parse_json_text(value: str, fallback: object) -> object:
    try:
        return json.loads(value or "")
    except (TypeError, ValueError, json.JSONDecodeError):
        return fallback


class ArchiveHandler(BaseHTTPRequestHandler):
    server_version = "QoyodArchive/2.0"

    def request_user(self, query: str = "", explicit: str = "") -> str:
        # Prefer session-based authenticated user when available
        try:
            cookie = SimpleCookie(self.headers.get("Cookie"))
            token = cookie.get(SESSION_COOKIE).value if cookie and cookie.get(
                SESSION_COOKIE) else ""
            if token:
                with connect_db() as conn:
                    user = session_user(conn, token)
                    if user:
                        return user.get("display_name") or user.get("username")
        except Exception:
            pass
        # fallback to previous behavior for legacy requests or initial setup
        header_user = unquote(
            (self.headers.get("X-Archive-User") or "").strip())
        if not query:
            query = urlparse(self.path).query
        query_user = ""
        if query:
            query_user = (parse_qs(query).get("user", [""])[0] or "").strip()
        return (header_user or explicit.strip() or query_user)[:100]

    def client_ip(self) -> str:
        return str(self.client_address[0] if self.client_address else "")

    def require_user(self, explicit: str = "") -> str:
        actor = self.request_user(explicit=explicit)
        if not actor:
            raise ValueError("اكتب اسم المستخدم قبل استخدام النظام.")
        return actor

    def record_activity(
        self,
        action: str,
        entity_type: str,
        entity_id: object,
        details: object = "",
        user_name: str = "",
    ) -> None:
        actor = self.require_user(user_name)
        with connect_db() as conn:
            audit(
                conn,
                action,
                entity_type,
                entity_id,
                actor,
                "",
                details,
                self.client_ip(),
            )
            conn.commit()

    def send_json(self, payload: object, status: int = 200) -> None:
        data = json_bytes(payload)
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_bytes(
        self,
        data: bytes,
        content_type: str,
        filename: str = "",
        status: int = 200,
    ) -> None:
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("X-Content-Type-Options", "nosniff")
        if filename:
            safe_name = filename.replace('"', "")
            self.send_header("Content-Disposition",
                             f'attachment; filename="{safe_name}"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def send_error_json(self, message: str, status: int = 400) -> None:
        self.send_json({"ok": False, "error": message}, status)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        try:
            if path == "/health":
                return self.handle_health()
            if path == "/api/meta":
                return self.handle_meta()
            if path == "/api/me":
                # return current session user if any
                cookie = SimpleCookie(self.headers.get("Cookie"))
                token = cookie.get(SESSION_COOKIE).value if cookie and cookie.get(
                    SESSION_COOKIE) else ""
                if not token:
                    return self.send_json({"ok": True, "user": None})
                with connect_db() as conn:
                    user = session_user(conn, token)
                return self.send_json({"ok": True, "user": user})
            if path == "/api/setup":
                with connect_db() as conn:
                    count = conn.execute(
                        "SELECT COUNT(*) FROM users").fetchone()[0]
                return self.send_json({"ok": True, "has_users": bool(count)})
            if path == "/api/settings":
                return self.handle_settings()
            if path == "/api/users":
                # list users (admin only)
                cookie = SimpleCookie(self.headers.get("Cookie"))
                token = cookie.get(SESSION_COOKIE).value if cookie and cookie.get(
                    SESSION_COOKIE) else ""
                with connect_db() as conn:
                    current = session_user(conn, token) if token else None
                    if not current or current.get("role") != "admin":
                        return self.send_json({"ok": False, "error": "فقط المدير يمكنه عرض المستخدمين."}, 403)
                    rows = conn.execute(
                        "SELECT id, username, display_name, role, is_active, created_at FROM users ORDER BY username").fetchall()
                    users = [dict(r) for r in rows]
                return self.send_json({"ok": True, "users": users})
            if path == "/api/stats":
                return self.handle_stats()
            if path == "/api/dashboard":
                return self.handle_dashboard()
            if path == "/api/excel/imports":
                return self.handle_imports()
            document_preview = re.fullmatch(
                r"/api/documents/(\d+)/preview", path)
            if document_preview:
                return self.handle_document_preview(int(document_preview.group(1)))
            if path == "/api/documents":
                return self.handle_documents(parsed.query)
            if path == "/api/audit":
                return self.handle_audit(parsed.query)
            if path.startswith("/api/reports/"):
                return self.handle_report(path.removeprefix("/api/reports/"), parsed.query)
            if path == "/api/entries":
                return self.handle_entries(parsed.query)
            if path.startswith("/api/entries/") and path.endswith("/cover-sheet"):
                trans_no = unquote(path.removeprefix(
                    "/api/entries/").removesuffix("/cover-sheet").strip("/"))
                return self.handle_cover_sheet(trans_no)
            if path.startswith("/api/entries/"):
                trans_no = unquote(path.removeprefix("/api/entries/"))
                return self.handle_entry_detail(trans_no)
            if path.startswith("/files/"):
                attachment_id = int(path.removeprefix("/files/"))
                return self.handle_file(attachment_id)
            return self.handle_static(path)
        except Exception as exc:
            LOGGER.exception("GET %s failed", self.path)
            return self.send_error_json(str(exc), 400 if isinstance(exc, ValueError) else 500)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        try:
            if path == "/api/login":
                payload = read_json(self)
                username = payload.get("username") or payload.get("user")
                password = payload.get("password")
                if not username or not password:
                    data = json_bytes(
                        {"ok": False, "error": "اسم المستخدم وكلمة المرور مطلوبان."})
                    self.send_response(400)
                    self.send_header(
                        "Content-Type", "application/json; charset=utf-8")
                    self.send_header("Content-Length", str(len(data)))
                    self.end_headers()
                    self.wfile.write(data)
                    return
                with connect_db() as conn:
                    user = authenticate(conn, username, password)
                    if not user:
                        data = json_bytes(
                            {"ok": False, "error": "اسم المستخدم أو كلمة المرور غير صحيحة."})
                        self.send_response(401)
                        self.send_header(
                            "Content-Type", "application/json; charset=utf-8")
                        self.send_header("Content-Length", str(len(data)))
                        self.end_headers()
                        self.wfile.write(data)
                        return
                    token = create_session(
                        conn, int(user["id"]), self.client_ip())
                    payload = {"ok": True, "user": user}
                    data = json_bytes(payload)
                    # set secure cookie (no SameSite considerations for local)
                    cookie = f"{SESSION_COOKIE}={token}; Path=/; HttpOnly"
                    self.send_response(200)
                    self.send_header(
                        "Content-Type", "application/json; charset=utf-8")
                    self.send_header("Set-Cookie", cookie)
                    self.send_header("Content-Length", str(len(data)))
                    self.end_headers()
                    self.wfile.write(data)
                    with connect_db() as conn2:
                        audit(conn2, "تسجيل دخول", "session", f"session:{user['username']}", user.get(
                            "display_name") or user.get("username"), "", "", self.client_ip())
                        conn2.commit()
                    return
            if path == "/api/logout":
                # revoke session cookie
                cookie = SimpleCookie(self.headers.get("Cookie"))
                token = cookie.get(SESSION_COOKIE).value if cookie and cookie.get(
                    SESSION_COOKIE) else ""
                if token:
                    with connect_db() as conn:
                        revoke_session(conn, token)
                        conn.commit()
                # expire cookie
                self.send_response(200)
                self.send_header(
                    "Content-Type", "application/json; charset=utf-8")
                self.send_header(
                    "Set-Cookie", f"{SESSION_COOKIE}=deleted; Path=/; HttpOnly; Expires=Thu, 01 Jan 1970 00:00:00 GMT")
                data = json_bytes({"ok": True})
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return
            if path in {"/api/import", "/api/excel/import"}:
                return self.handle_import()
            if path == "/api/users":
                payload = read_json(self)
                username = payload.get("username")
                display_name = payload.get("display_name") or payload.get(
                    "displayName") or username
                password = payload.get("password")
                role = payload.get("role") or "user"
                with connect_db() as conn:
                    users_count = conn.execute(
                        "SELECT COUNT(*) FROM users").fetchone()[0]
                    # allow creating first admin without authentication
                    if users_count == 0:
                        new = create_user(
                            conn, username, display_name, password, role="admin")
                        conn.commit()
                        return self.send_json({"ok": True, "user": new})
                    # otherwise require current session to be admin
                    cookie = SimpleCookie(self.headers.get("Cookie"))
                    token = cookie.get(SESSION_COOKIE).value if cookie and cookie.get(
                        SESSION_COOKIE) else ""
                    current = None
                    if token:
                        current = session_user(conn, token)
                    if not current or current.get("role") != "admin":
                        return self.send_json({"ok": False, "error": "فقط المدير يمكنه إنشاء مستخدمين."}, 403)
                    new = create_user(conn, username, display_name, password,
                                      role="admin" if role == "admin" else "user")
                    conn.commit()
                    return self.send_json({"ok": True, "user": new})
            if path == "/api/paper-register/import":
                return self.handle_paper_import()
            if path == "/api/excel/preview":
                return self.handle_excel_preview()
            if path == "/api/entries":
                payload = read_json(self)
                actor = self.require_user(str(payload.get("user_name") or ""))
                with connect_db() as conn:
                    trans_no = str(payload.get("trans_no") or "").strip()
                    if not conn.execute("SELECT 1 FROM entries WHERE trans_no = ?", (trans_no,)).fetchone():
                        raise ValueError(
                            "لا يمكن إنشاء قيد يدوي. حدّث كشف الأرشيف المعتمد أولًا.")
                    upsert_entry(conn, payload)
                    audit(
                        conn,
                        "حفظ بيانات قيد",
                        "entry",
                        trans_no,
                        actor,
                        "",
                        payload,
                        self.client_ip(),
                    )
                    conn.commit()
                return self.send_json({"ok": True, "trans_no": trans_no})
            if path == "/api/activity":
                return self.handle_activity()
            if path == "/api/documents/upload":
                return self.handle_document_upload()
            workflow_match = re.fullmatch(
                r"/api/(?:journals|entries)/([^/]+)/(archive|unarchive)",
                path,
            )
            if workflow_match:
                return self.handle_workflow(unquote(workflow_match.group(1)), workflow_match.group(2))
            if path.startswith("/api/entries/") and path.endswith("/attachments"):
                trans_no = unquote(path.removeprefix(
                    "/api/entries/").removesuffix("/attachments").strip("/"))
                return self.handle_attachment_upload(trans_no)
            if path.startswith("/api/entries/") and path.endswith("/paper-status"):
                trans_no = unquote(path.removeprefix(
                    "/api/entries/").removesuffix("/paper-status").strip("/"))
                return self.handle_paper_status(trans_no)
            if path.startswith("/api/entries/") and path.endswith("/from-url"):
                trans_no = unquote(path.removeprefix(
                    "/api/entries/").removesuffix("/from-url").strip("/"))
                return self.handle_attachment_from_url(trans_no)
            if path.startswith("/api/entries/") and path.endswith("/notes"):
                trans_no = unquote(path.removeprefix(
                    "/api/entries/").removesuffix("/notes").strip("/"))
                return self.handle_note(trans_no)
            return self.send_error_json("مسار غير معروف.", 404)
        except Exception as exc:
            LOGGER.exception("POST %s failed", self.path)
            return self.send_error_json(str(exc), 400)

    def do_PUT(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        try:
            if path.startswith("/api/entries/"):
                trans_no = unquote(path.removeprefix("/api/entries/"))
                payload = read_json(self)
                payload["trans_no"] = trans_no
                actor = self.require_user(str(payload.get("user_name") or ""))
                with connect_db() as conn:
                    old = conn.execute(
                        "SELECT * FROM entries WHERE trans_no = ?", (trans_no,)).fetchone()
                    if not old:
                        raise ValueError(
                            "رقم القيد غير موجود في كشف الأرشيف المعتمد.")
                    upsert_entry(conn, payload)
                    audit(
                        conn,
                        "تحديث قيد",
                        "entry",
                        trans_no,
                        actor,
                        dict(old) if old else "",
                        payload,
                        self.client_ip(),
                    )
                    conn.commit()
                return self.send_json({"ok": True, "trans_no": trans_no})
            return self.send_error_json("مسار غير معروف.", 404)
        except Exception as exc:
            LOGGER.exception("PUT %s failed", self.path)
            return self.send_error_json(str(exc), 400)

    def handle_static(self, path: str) -> None:
        if path in {"/", ""}:
            file_path = STATIC_DIR / "index.html"
        else:
            clean = path.lstrip("/")
            file_path = (STATIC_DIR / clean).resolve()
            if STATIC_DIR.resolve() not in file_path.parents and file_path != STATIC_DIR.resolve():
                return self.send_error(403)
        if not file_path.exists() or not file_path.is_file():
            return self.send_error(404)
        mime = mimetypes.guess_type(file_path.name)[
            0] or "application/octet-stream"
        data = file_path.read_bytes()
        self.send_response(200)
        self.send_header(
            "Content-Type", f"{mime}; charset=utf-8" if mime.startswith("text/") else mime)
        self.send_header("Cache-Control", "no-store")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def handle_health(self) -> None:
        with connect_db() as conn:
            conn.execute("SELECT 1").fetchone()
        self.send_json(
            {
                "ok": True,
                "status": "healthy",
                "version": "2.0",
                "database": str(DB_PATH),
                "ocr_provider": SETTINGS.ocr_provider,
                "time": now_iso(),
            }
        )

    def handle_meta(self) -> None:
        with connect_db() as conn:
            rows = conn.execute(
                """
                SELECT DISTINCT accountant
                FROM entries
                WHERE TRIM(COALESCE(accountant, '')) != ''
                ORDER BY accountant
                """
            ).fetchall()
        self.send_json(
            {
                "ok": True,
                "statuses": STATUSES,
                "accountants": [row["accountant"] for row in rows],
            }
        )

    def handle_settings(self) -> None:
        self.send_json(
            {
                "ok": True,
                "active_ocr_provider": "none",
                "ocr_languages": SETTINGS.ocr_languages,
                "providers": {},
                "manual_review_available": True,
                "data_dir": str(DATA_DIR),
            }
        )

    def handle_imports(self) -> None:
        with connect_db() as conn:
            rows = conn.execute(
                "SELECT * FROM excel_imports ORDER BY imported_at DESC LIMIT 100"
            ).fetchall()
        self.send_json({"ok": True, "imports": [dict(row) for row in rows]})

    def handle_documents(self, query: str) -> None:
        params = parse_qs(query)
        trans_no = (params.get("trans_no", [""])[0] or "").strip()
        clauses = []
        args: list[object] = []
        if trans_no:
            clauses.append("a.trans_no = ?")
            args.append(trans_no)
        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        with connect_db() as conn:
            rows = conn.execute(
                f"""
                SELECT a.*, e.entry_date, e.amount, e.project, e.name
                FROM attachments a
                JOIN entries e ON e.trans_no = a.trans_no
                {where}
                ORDER BY a.created_at DESC
                LIMIT 200
                """,
                args,
            ).fetchall()
        documents = []
        for row in rows:
            item = dict(row)
            item["extraction"] = parse_json_text(
                item.get("extraction_json", ""), {})
            item.pop("extraction_json", None)
            documents.append(item)
        self.send_json({"ok": True, "documents": documents})

    def handle_document_preview(self, attachment_id: int) -> None:
        with connect_db() as conn:
            row = conn.execute(
                "SELECT * FROM attachments WHERE id = ?", (attachment_id,)).fetchone()
        if not row:
            return self.send_error_json("المرفق غير موجود.", 404)
        path = attachment_path(row)
        if not path.exists():
            return self.send_error_json("ملف المرفق غير موجود على الجهاز.", 404)
        self.record_activity(
            "معاينة مرفق",
            "attachment",
            attachment_id,
            {"trans_no": row["trans_no"], "file_name": row["original_name"]},
        )

        suffix = path.suffix.lower()
        mime = row["mime"] or mimetypes.guess_type(
            path.name)[0] or "application/octet-stream"
        if mime == "application/pdf" or suffix == ".pdf":
            return self.send_bytes(path.read_bytes(), "application/pdf")
        if suffix in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
            return self.send_bytes(path.read_bytes(), mime)

        try:
            from PIL import Image
            from pillow_heif import register_heif_opener

            register_heif_opener()
            image = Image.open(path)
            image.thumbnail((2200, 2200))
            if image.mode not in {"RGB", "L"}:
                image = image.convert("RGB")
            output = io.BytesIO()
            image.save(output, format="JPEG", quality=90, optimize=True)
            return self.send_bytes(output.getvalue(), "image/jpeg")
        except Exception as exc:
            LOGGER.exception(
                "Preview generation failed for attachment %s", attachment_id)
            raise RuntimeError(f"تعذر إنشاء معاينة لهذا الملف: {exc}") from exc

    def handle_audit(self, query: str) -> None:
        self.require_user()
        params = parse_qs(query)
        entity_id = (params.get("entity_id", [""])[0] or "").strip()
        user_name = (params.get("user_name", [""])[0] or "").strip()
        action = (params.get("action", [""])[0] or "").strip()
        text = (params.get("q", [""])[0] or "").strip()
        clauses = []
        args: list[object] = []
        if entity_id:
            clauses.append("entity_id = ?")
            args.append(entity_id)
        if user_name:
            clauses.append("user_name = ?")
            args.append(user_name)
        if action:
            clauses.append("action = ?")
            args.append(action)
        if text:
            clauses.append(
                "(action LIKE ? OR entity_id LIKE ? OR new_value LIKE ?)")
            like = f"%{text}%"
            args.extend([like, like, like])
        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        with connect_db() as conn:
            rows = conn.execute(
                f"SELECT * FROM audit_logs {where} ORDER BY created_at DESC LIMIT 300",
                args,
            ).fetchall()
            users = conn.execute(
                "SELECT DISTINCT user_name FROM audit_logs WHERE user_name <> '' ORDER BY user_name"
            ).fetchall()
            actions = conn.execute(
                "SELECT DISTINCT action FROM audit_logs ORDER BY action"
            ).fetchall()
        self.send_json(
            {
                "ok": True,
                "logs": [dict(row) for row in rows],
                "users": [row[0] for row in users],
                "actions": [row[0] for row in actions],
            }
        )

    def handle_dashboard(self) -> None:
        self.require_user()
        with connect_db() as conn:
            archived = conn.execute(
                "SELECT COUNT(*) FROM entries WHERE status = 'مؤرشف'"
            ).fetchone()[0]
            pending = conn.execute(
                """
                SELECT COUNT(*) FROM entries
                WHERE status IN ('بانتظار المرفقات', 'تم رفع المرفقات')
                """
            ).fetchone()[0]
            with_attachments = conn.execute(
                """
                SELECT COUNT(DISTINCT e.trans_no) FROM entries e
                JOIN attachments a ON a.trans_no = e.trans_no
                """
            ).fetchone()[0]
            missing = conn.execute(
                """
                SELECT COUNT(*) FROM entries e
                LEFT JOIN attachments a ON a.trans_no = e.trans_no
                WHERE a.id IS NULL
                """
            ).fetchone()[0]
            today_actions = conn.execute(
                "SELECT COUNT(*) FROM audit_logs WHERE date(created_at) = date('now', 'localtime')"
            ).fetchone()[0]
            queue = conn.execute(
                """
                SELECT e.trans_no, e.entry_date, e.name, e.accountant, e.amount,
                       e.status, COUNT(a.id) attachments_count
                FROM entries e
                LEFT JOIN attachments a ON a.trans_no = e.trans_no
                WHERE e.status != 'مؤرشف'
                GROUP BY e.trans_no
                ORDER BY e.updated_at DESC
                LIMIT 100
                """
            ).fetchall()
            recent = conn.execute(
                "SELECT * FROM audit_logs ORDER BY created_at DESC LIMIT 12"
            ).fetchall()
        self.send_json(
            {
                "ok": True,
                "archived": archived,
                "pending": pending,
                "with_attachments": with_attachments,
                "missing": missing,
                "today_actions": today_actions,
                "queue": [dict(row) for row in queue],
                "recent": [dict(row) for row in recent],
            }
        )

    def handle_activity(self) -> None:
        payload = read_json(self)
        actor = self.request_user(explicit=str(payload.get("user_name") or ""))
        if not actor:
            raise ValueError("اكتب اسم المستخدم أولًا.")
        action = str(payload.get("action") or "").strip()[:120]
        entity_type = str(payload.get("entity_type") or "screen").strip()[:60]
        entity_id = str(payload.get("entity_id") or "").strip()[:160]
        if not action:
            raise ValueError("نوع الحركة مطلوب.")
        self.record_activity(
            action, entity_type, entity_id or "-", payload.get("details") or "", actor)
        self.send_json({"ok": True})

    def _report_rows(self, report_name: str) -> tuple[list[str], list[sqlite3.Row], str]:
        reports = {
            "daily": (
                """
                SELECT e.trans_no, e.entry_date, e.entry_type, e.num, e.name, e.description,
                       e.amount, e.paper_received, e.status, COUNT(a.id) attachments_count
                FROM entries e LEFT JOIN attachments a ON a.trans_no = e.trans_no
                WHERE date(e.updated_at) = date('now', 'localtime')
                GROUP BY e.trans_no ORDER BY e.updated_at DESC
                """,
                "التقرير اليومي",
            ),
            "missing": (
                """
                SELECT e.trans_no, e.entry_date, e.entry_type, e.num, e.name, e.description,
                       e.amount, e.status
                FROM entries e
                WHERE COALESCE(e.paper_received, 0) = 0
                ORDER BY e.trans_no
                """,
                "قيود QuickBooks التي لم تصل ورقيًا",
            ),
            "paper-received": (
                """
                SELECT e.trans_no, e.entry_date, e.entry_type, e.num, e.name, e.description,
                       e.amount, e.paper_received_at, e.paper_received_by
                FROM entries e
                WHERE COALESCE(e.paper_received, 0) = 1
                ORDER BY e.trans_no
                """,
                "القيود الموجودة ورقيًا",
            ),
            "attachments-missing": (
                """
                SELECT e.trans_no, e.entry_date, e.entry_type, e.num, e.name, e.description,
                       e.amount, e.paper_received, e.status
                FROM entries e LEFT JOIN attachments a ON a.trans_no = e.trans_no
                WHERE a.id IS NULL ORDER BY e.trans_no
                """,
                "القيود بدون مرفقات إلكترونية",
            ),
            "duplicates": (
                """
                SELECT a.id, a.trans_no, a.original_name, a.duplicate_of, a.created_at
                FROM attachments a WHERE a.duplicate_of IS NOT NULL ORDER BY a.created_at DESC
                """,
                "المرفقات المكررة",
            ),
            "archived": (
                """
                SELECT e.trans_no, e.entry_date, e.entry_type, e.num, e.name, e.description,
                       e.amount, e.accountant, e.archive_date
                FROM entries e
                WHERE e.status = 'مؤرشف'
                ORDER BY e.archive_date DESC, e.trans_no
                """,
                "القيود المؤرشفة",
            ),
            "activity-log": (
                """
                SELECT user_name, action, entity_type, entity_id, ip_address,
                       old_value, new_value, created_at
                FROM audit_logs
                ORDER BY created_at DESC
                LIMIT 10000
                """,
                "سجل نشاط المستخدمين",
            ),
        }
        sql, title = reports.get(report_name, reports["daily"])
        with connect_db() as conn:
            rows = conn.execute(sql).fetchall()
        headers = list(rows[0].keys()) if rows else []
        return headers, rows, title

    def handle_report(self, report_name: str, query: str) -> None:
        params = parse_qs(query)
        output_format = (params.get("format", ["xlsx"])[0] or "xlsx").lower()
        headers, rows, title = self._report_rows(report_name)
        self.record_activity(
            "فتح تقرير" if output_format == "json" else "تحميل تقرير",
            "report",
            report_name,
            {"format": output_format, "rows": len(rows), "title": title},
        )
        if output_format == "json":
            return self.send_json({"ok": True, "title": title, "rows": [dict(row) for row in rows]})
        if output_format == "csv":
            stream = io.StringIO()
            writer = csv.writer(stream)
            writer.writerow(headers)
            writer.writerows([[row[header] for header in headers]
                             for row in rows])
            data = ("\ufeff" + stream.getvalue()).encode("utf-8")
            return self.send_bytes(data, "text/csv; charset=utf-8", f"{report_name}.csv")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise RuntimeError(
                "تصدير Excel يحتاج openpyxl. شغّل setup.bat.") from exc
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title[:31]
        sheet.sheet_view.rightToLeft = True
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="176B91")
            cell.alignment = Alignment(horizontal="right")
        for row in rows:
            sheet.append([row[header] for header in headers])
        for column in sheet.columns:
            width = min(40, max(12, max(len(str(cell.value or ""))
                        for cell in column) + 2))
            sheet.column_dimensions[column[0].column_letter].width = width
        output = io.BytesIO()
        workbook.save(output)
        self.send_bytes(
            output.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"{report_name}.xlsx",
        )

    def handle_cover_sheet(self, trans_no: str) -> None:
        with connect_db() as conn:
            entry = conn.execute(
                "SELECT * FROM entries WHERE trans_no = ?", (trans_no,)).fetchone()
        if not entry:
            return self.send_error_json("القيد غير موجود.", 404)
        self.record_activity("فتح غلاف QR", "entry", trans_no)
        try:
            import base64

            import qrcode
        except ImportError as exc:
            raise RuntimeError(
                "توليد QR يحتاج مكتبة qrcode. شغّل setup.bat.") from exc
        image = qrcode.make(f"QOYOD:{trans_no}")
        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
        qr = base64.b64encode(buffer.getvalue()).decode("ascii")
        html = f"""<!doctype html>
<html lang="ar" dir="rtl"><head><meta charset="utf-8"><title>غلاف القيد {trans_no}</title>
<style>
body{{font-family:Tahoma,Arial;margin:0;padding:30px;color:#152020}}
.sheet{{border:4px solid #176b91;padding:32px;min-height:900px}}
.head{{display:flex;justify-content:space-between;align-items:center;border-bottom:2px solid #152020;padding-bottom:20px}}
.qr{{width:210px;height:210px}} h1{{font-size:42px;margin:0}} .number{{font-size:72px;font-weight:900;color:#0e506d}}
dl{{display:grid;grid-template-columns:1fr 1fr;gap:18px;margin-top:30px}} div{{border:1px solid #aaa;padding:16px}}
dt{{color:#667174;font-weight:bold}} dd{{font-size:22px;font-weight:bold;margin:8px 0 0}}
@media print{{button{{display:none}} body{{padding:0}} .sheet{{border-width:3px}}}}
</style></head><body><button onclick="print()">طباعة</button><section class="sheet">
<div class="head"><div><h1>غلاف مرفقات القيد</h1><div class="number">{trans_no}</div></div>
<img class="qr" src="data:image/png;base64,{qr}" alt="QR"></div>
<dl>
<div><dt>التاريخ</dt><dd>{entry['entry_date'] or '-'}</dd></div>
<div><dt>النوع</dt><dd>{entry['entry_type'] or '-'}</dd></div>
<div><dt>المشروع</dt><dd>{entry['project'] or '-'}</dd></div>
<div><dt>المبلغ</dt><dd>{entry['amount'] or 0:,.2f} EGP</dd></div>
<div><dt>المحاسب</dt><dd>{entry['accountant'] or '-'}</dd></div>
<div><dt>الجهة</dt><dd>{entry['supplier'] or entry['client'] or entry['name'] or '-'}</dd></div>
</dl></section></body></html>"""
        self.send_bytes(html.encode("utf-8"), "text/html; charset=utf-8")

    def handle_stats(self) -> None:
        with connect_db() as conn:
            total = conn.execute("SELECT COUNT(*) FROM entries").fetchone()[0]
            missing = conn.execute(
                """
                SELECT COUNT(*)
                FROM entries e
                LEFT JOIN attachments a ON a.trans_no = e.trans_no
                WHERE a.id IS NULL
                """
            ).fetchone()[0]
            attachments = conn.execute(
                "SELECT COUNT(*), COALESCE(SUM(size), 0) FROM attachments").fetchone()
            status_rows = conn.execute(
                "SELECT status, COUNT(*) c FROM entries GROUP BY status").fetchall()
            duplicates = conn.execute(
                "SELECT COUNT(*) FROM attachments WHERE duplicate_of IS NOT NULL"
            ).fetchone()[0]
            archived = conn.execute(
                "SELECT COUNT(*) FROM entries WHERE status = 'مؤرشف'"
            ).fetchone()[0]
            scanned = conn.execute(
                "SELECT COUNT(DISTINCT trans_no) FROM attachments").fetchone()[0]
            paper_received = conn.execute(
                "SELECT COUNT(*) FROM entries WHERE COALESCE(paper_received, 0) = 1"
            ).fetchone()[0]
            paper_missing = total - paper_received
        self.send_json(
            {
                "ok": True,
                "total": total,
                "missing": missing,
                "attachments_count": attachments[0],
                "attachments_size": attachments[1],
                "scanned": scanned,
                "archived": archived,
                "duplicates": duplicates,
                "paper_received": paper_received,
                "paper_missing": paper_missing,
                "by_status": {row["status"]: row["c"] for row in status_rows},
            }
        )

    def handle_entries(self, query: str) -> None:
        params = parse_qs(query)
        q = (params.get("q", [""])[0] or "").strip()
        status = (params.get("status", [""])[0] or "").strip()
        accountant = (params.get("accountant", [""])[0] or "").strip()
        missing = (params.get("missing", [""])[0] or "") == "1"
        paper = (params.get("paper", [""])[0] or "").strip()
        clauses = []
        args: list[object] = []
        if q:
            like = f"%{q}%"
            clauses.append(
                "(e.trans_no LIKE ? OR e.name LIKE ? OR e.num LIKE ? OR "
                "e.description LIKE ? OR e.entry_type LIKE ? OR e.accountant LIKE ?)"
            )
            args.extend([like, like, like, like, like, like])
        if status:
            clauses.append("e.status = ?")
            args.append(status)
        if accountant:
            clauses.append("e.accountant = ?")
            args.append(accountant)
        if missing:
            clauses.append("a.id IS NULL")
        if paper == "present":
            clauses.append("COALESCE(e.paper_received, 0) = 1")
        elif paper == "missing":
            clauses.append("COALESCE(e.paper_received, 0) = 0")
        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        with connect_db() as conn:
            rows = conn.execute(
                f"""
                SELECT e.*, COUNT(a.id) attachments_count
                FROM entries e
                LEFT JOIN attachments a ON a.trans_no = e.trans_no
                {where}
                GROUP BY e.trans_no
                ORDER BY e.updated_at DESC, e.trans_no DESC
                LIMIT 200
                """,
                args,
            ).fetchall()
        self.record_activity(
            "بحث في القيود",
            "search",
            q or "الكل",
            {
                "query": q,
                "status": status,
                "accountant": accountant,
                "missing_attachments": missing,
                "results": len(rows),
            },
        )
        self.send_json(
            {"ok": True, "entries": [row_to_entry(row) for row in rows]})

    def handle_entry_detail(self, trans_no: str) -> None:
        with connect_db() as conn:
            row = conn.execute(
                """
                SELECT e.*, COUNT(a.id) attachments_count
                FROM entries e
                LEFT JOIN attachments a ON a.trans_no = e.trans_no
                WHERE e.trans_no = ?
                GROUP BY e.trans_no
                """,
                (trans_no,),
            ).fetchone()
            if not row:
                return self.send_error_json("القيد غير موجود.", 404)
            attachments = conn.execute(
                "SELECT * FROM attachments WHERE trans_no = ? ORDER BY created_at DESC",
                (trans_no,),
            ).fetchall()
            notes = conn.execute(
                "SELECT * FROM notes WHERE trans_no = ? ORDER BY created_at DESC",
                (trans_no,),
            ).fetchall()
            audit_rows = conn.execute(
                "SELECT * FROM audit_logs WHERE entity_id = ? ORDER BY created_at DESC LIMIT 50",
                (trans_no,),
            ).fetchall()
        self.record_activity(
            "فتح القيد",
            "entry",
            trans_no,
            {"attachments": len(attachments), "status": row["status"]},
        )
        self.send_json(
            {
                "ok": True,
                "entry": row_to_entry(row),
                "attachments": [dict(item) for item in attachments],
                "notes": [dict(item) for item in notes],
                "audit": [dict(item) for item in audit_rows],
            }
        )

    def handle_excel_preview(self) -> None:
        fields, files = parse_multipart(self)
        if not files:
            raise ValueError("اختر ملف Excel أو CSV.")
        headers = import_headers(files[0]["filename"], files[0]["content"])
        mode = (fields.get("mode", [""])[0] or "").strip()
        detected_mapping = suggested_mapping(headers)
        self.record_activity(
            "معاينة ملف كشف الأرشيف",
            "excel",
            files[0]["filename"],
            {"headers": len(headers)},
        )
        if mode == "paper":
            visible_keys = ["trans_no"]
        else:
            visible_keys = [
                key for key in QUICKBOOKS_MAPPING_FIELDS if key in detected_mapping
            ]
        self.send_json(
            {
                "ok": True,
                "filename": files[0]["filename"],
                "headers": headers,
                "suggested_mapping": {
                    key: detected_mapping[key]
                    for key in visible_keys
                    if key in detected_mapping
                },
                "fields": [
                    {
                        "key": key,
                        "label": detected_mapping.get(key, IMPORT_ALIASES[key][0]),
                    }
                    for key in visible_keys
                ],
            }
        )

    def handle_import(self) -> None:
        fields, files = parse_multipart(self)
        if not files:
            raise ValueError("اختار ملف Excel أو CSV.")
        mapping = {}
        if fields.get("mapping"):
            parsed_mapping = parse_json_text(fields["mapping"][0], {})
            if isinstance(parsed_mapping, dict):
                mapping = {str(key): str(value)
                           for key, value in parsed_mapping.items() if value}
        imported_by = self.request_user(
            explicit=(fields.get("imported_by", [""])[0] or "").strip()
        )
        authoritative_archive = (fields.get(
            "authoritative_archive", [""])[0] or "") == "1"
        if authoritative_archive and not imported_by:
            raise ValueError("اكتب اسم منفذ تحديث كشف الأرشيف.")
        imported = 0
        updated = 0
        records = parse_import_rows(
            files[0]["filename"], files[0]["content"], mapping)
        if not records:
            raise ValueError("لم يتم العثور على قيود صالحة داخل الملف.")
        duplicate_rows = len(records) - \
            len({record["trans_no"] for record in records})
        missing_amount = sum(1 for record in records if not record["amount"])
        missing_date = sum(1 for record in records if not record["entry_date"])
        imported_at = now_iso()
        paper_present = sum(
            1 for record in records if record.get("paper_received") is True)
        paper_missing = sum(1 for record in records if record.get(
            "paper_received") is False)
        removed = 0
        if authoritative_archive:
            backup_dir = DATA_DIR / "backups"
            backup_dir.mkdir(parents=True, exist_ok=True)
            backup_path = backup_dir / \
                f"archive-before-register-{datetime.now():%Y%m%d-%H%M%S}.db"
            backup_database(backup_path)
        with connect_db() as conn:
            source_ids = {record["trans_no"] for record in records}
            if authoritative_archive:
                existing_ids = {row[0] for row in conn.execute(
                    "SELECT trans_no FROM entries")}
                excluded_ids = existing_ids - source_ids
                if excluded_ids:
                    conn.execute(
                        "CREATE TEMP TABLE import_archive_ids (trans_no TEXT PRIMARY KEY)")
                    conn.executemany(
                        "INSERT INTO import_archive_ids (trans_no) VALUES (?)",
                        ((trans_no,) for trans_no in source_ids),
                    )
                    attached_excluded = conn.execute(
                        """
                        SELECT COUNT(*)
                        FROM attachments a
                        WHERE NOT EXISTS (
                            SELECT 1 FROM import_archive_ids i WHERE i.trans_no = a.trans_no
                        )
                        """
                    ).fetchone()[0]
                    if attached_excluded:
                        raise ValueError(
                            f"تعذر اعتماد الكشف: يوجد {attached_excluded} مرفق مرتبط بقيود غير موجودة في الملف."
                        )
                    removed = conn.execute(
                        """
                        SELECT COUNT(*) FROM entries e
                        WHERE NOT EXISTS (
                            SELECT 1 FROM import_archive_ids i WHERE i.trans_no = e.trans_no
                        )
                        """
                    ).fetchone()[0]
                    conn.execute(
                        """
                        DELETE FROM entries
                        WHERE NOT EXISTS (
                            SELECT 1 FROM import_archive_ids i WHERE i.trans_no = entries.trans_no
                        )
                        """
                    )
            suffix = Path(files[0]["filename"]).suffix.lower()
            stored_name = f"{datetime.now():%Y%m%d-%H%M%S}_{uuid.uuid4().hex[:8]}{suffix}"
            (SETTINGS.excel_dir / stored_name).write_bytes(files[0]["content"])
            for record in records:
                exists = conn.execute(
                    "SELECT 1 FROM entries WHERE trans_no = ?", (record["trans_no"],)).fetchone()
                upsert_entry(conn, record)
                if authoritative_archive or record.get("paper_received") is not None:
                    present = True if authoritative_archive else bool(
                        record["paper_received"])
                    conn.execute(
                        """
                        UPDATE entries
                        SET paper_received = ?, paper_received_at = ?,
                            paper_received_by = ?, paper_source = ?, updated_at = ?
                        WHERE trans_no = ?
                        """,
                        (
                            int(present),
                            imported_at if present else "",
                            imported_by,
                            stored_name,
                            imported_at,
                            record["trans_no"],
                        ),
                    )
                if exists:
                    updated += 1
                else:
                    imported += 1
            cursor = conn.execute(
                """
                INSERT INTO excel_imports (
                    file_name, stored_name, imported_by, imported_rows, updated_rows, mapping_json, imported_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    files[0]["filename"],
                    stored_name,
                    imported_by,
                    imported,
                    updated,
                    json.dumps(mapping, ensure_ascii=False),
                    imported_at,
                ),
            )
            audit(
                conn,
                "استيراد Excel",
                "excel_import",
                cursor.lastrowid,
                imported_by,
                "",
                {
                    "file": files[0]["filename"],
                    "new": imported,
                    "updated": updated,
                    "removed": removed,
                    "authoritative_archive": authoritative_archive,
                    "paper_present": paper_present,
                    "paper_missing": paper_missing,
                },
                self.client_ip(),
            )
            conn.commit()
        self.send_json(
            {
                "ok": True,
                "imported": imported,
                "updated": updated,
                "total_rows": len(records),
                "filename": files[0]["filename"],
                "duplicate_rows": duplicate_rows,
                "missing_amount": missing_amount,
                "missing_date": missing_date,
                "paper_present": paper_present,
                "paper_missing": paper_missing,
                "removed": removed,
                "message": (
                    f"تم اعتماد كشف الأرشيف: {len(records)} قيد، وإضافة {imported}، "
                    f"وتحديث {updated}، واستبعاد {removed}."
                    if authoritative_archive
                    else f"تم استيراد {imported} قيد جديد وتحديث {updated} قيد."
                ),
            }
        )

    def handle_paper_import(self) -> None:
        fields, files = parse_multipart(self)
        if not files:
            raise ValueError("اختر كشف القيود الموجودة ورقيًا.")
        imported_by = self.request_user(
            explicit=(fields.get("imported_by", [""])[0] or "").strip()
        )
        if not imported_by:
            raise ValueError("اكتب اسم منفذ العملية.")
        mapping = {}
        if fields.get("mapping"):
            parsed_mapping = parse_json_text(fields["mapping"][0], {})
            if isinstance(parsed_mapping, dict):
                mapping = {str(key): str(value)
                           for key, value in parsed_mapping.items() if value}
        trans_numbers = parse_paper_register_rows(
            files[0]["filename"],
            files[0]["content"],
            mapping,
        )
        if not trans_numbers:
            raise ValueError("لم يتم العثور على أرقام قيود داخل الملف.")
        suffix = Path(files[0]["filename"]).suffix.lower()
        stored_name = f"paper-{datetime.now():%Y%m%d-%H%M%S}_{uuid.uuid4().hex[:8]}{suffix}"
        (SETTINGS.excel_dir / stored_name).write_bytes(files[0]["content"])
        current = now_iso()
        matched = 0
        unmatched: list[str] = []
        with connect_db() as conn:
            conn.execute(
                """
                UPDATE entries
                SET paper_received = 0, paper_received_at = '',
                    paper_received_by = '', paper_source = ''
                """
            )
            for trans_no in trans_numbers:
                cursor = conn.execute(
                    """
                    UPDATE entries
                    SET paper_received = 1, paper_received_at = ?,
                        paper_received_by = ?, paper_source = ?, updated_at = ?
                    WHERE trans_no = ?
                    """,
                    (current, imported_by, stored_name, current, trans_no),
                )
                if cursor.rowcount:
                    matched += 1
                else:
                    unmatched.append(trans_no)
            cursor = conn.execute(
                """
                INSERT INTO paper_imports (
                    file_name, stored_name, imported_by, total_rows,
                    matched_rows, unmatched_rows, imported_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    files[0]["filename"],
                    stored_name,
                    imported_by,
                    len(trans_numbers),
                    matched,
                    len(unmatched),
                    current,
                ),
            )
            audit(
                conn,
                "استيراد كشف الورق",
                "paper_import",
                cursor.lastrowid,
                imported_by,
                "",
                {
                    "file": files[0]["filename"],
                    "matched": matched,
                    "unmatched": len(unmatched),
                },
                self.client_ip(),
            )
            conn.commit()
        self.send_json(
            {
                "ok": True,
                "total": len(trans_numbers),
                "matched": matched,
                "unmatched": len(unmatched),
                "unmatched_samples": unmatched[:20],
                "message": f"تم تأكيد وجود {matched} قيد ورقيًا.",
            }
        )

    def handle_document_upload(self) -> None:
        fields, files = parse_multipart(self)
        trans_no = (fields.get("trans_no", [""])[0] or "").strip()
        uploaded_by = self.request_user(
            explicit=(fields.get("uploaded_by", [""])[0] or "").strip()
        )
        if not uploaded_by:
            raise ValueError("اكتب اسم منفذ العملية الذي رفع المرفقات.")
        if not trans_no:
            raise ValueError("رقم القيد مطلوب لربط المستند.")
        if not files:
            raise ValueError("اختر ملف PDF أو صورة.")
        ids = []
        with connect_db() as conn:
            for item in files:
                ids.append(
                    store_attachment(
                        conn,
                        trans_no,
                        item["filename"],
                        item["content"],
                        item["mime"],
                        source="upload",
                        uploaded_by=uploaded_by,
                        ip_address=self.client_ip(),
                    )
                )
            conn.commit()
        self.send_json({"ok": True, "document_ids": ids,
                       "count": len(ids), "trans_no": trans_no})

    def handle_workflow(self, trans_no: str, action: str) -> None:
        payload = read_json(self)
        user_name = self.request_user(
            explicit=str(payload.get("user_name") or ""))
        if not user_name:
            raise ValueError("اكتب اسم منفذ الإجراء قبل تحديث حالة القيد.")
        new_status = WORKFLOW_ACTIONS[action]
        with connect_db() as conn:
            row = conn.execute(
                "SELECT * FROM entries WHERE trans_no = ?", (trans_no,)).fetchone()
            if not row:
                return self.send_error_json("القيد غير موجود.", 404)
            archive_date = now_iso(
            ) if action == "archive" else row["archive_date"]
            conn.execute(
                """
                UPDATE entries
                SET status = ?, reviewer_status = ?, archive_date = ?, updated_at = ?
                WHERE trans_no = ?
                """,
                (new_status, action, archive_date, now_iso(), trans_no),
            )
            audit(
                conn,
                f"إجراء سير العمل: {action}",
                "entry",
                trans_no,
                user_name,
                row["status"],
                new_status,
                self.client_ip(),
            )
            conn.commit()
        self.send_json(
            {"ok": True, "trans_no": trans_no, "status": new_status})

    def handle_attachment_upload(self, trans_no: str) -> None:
        fields, files = parse_multipart(self)
        uploaded_by = self.request_user(
            explicit=(fields.get("uploaded_by", [""])[0] or "").strip()
        )
        if not uploaded_by:
            raise ValueError("اكتب اسم منفذ العملية الذي رفع المرفقات.")
        if not files:
            raise ValueError("اختار ملف PDF أو صورة.")
        ids = []
        with connect_db() as conn:
            for item in files:
                ids.append(
                    store_attachment(
                        conn,
                        trans_no,
                        item["filename"],
                        item["content"],
                        item["mime"],
                        source="upload",
                        uploaded_by=uploaded_by,
                        ip_address=self.client_ip(),
                    )
                )
            conn.commit()
        self.send_json({"ok": True, "attachment_ids": ids, "count": len(ids)})

    def handle_paper_status(self, trans_no: str) -> None:
        payload = read_json(self)
        user_name = self.request_user(
            explicit=str(payload.get("user_name") or ""))
        present = payload.get("present")
        if not user_name:
            raise ValueError("اكتب اسم منفذ العملية.")
        if not isinstance(present, bool):
            raise ValueError("حالة وجود الورق غير صحيحة.")
        with connect_db() as conn:
            row = conn.execute(
                "SELECT paper_received FROM entries WHERE trans_no = ?",
                (trans_no,),
            ).fetchone()
            if not row:
                return self.send_error_json("القيد غير موجود.", 404)
            current = now_iso()
            conn.execute(
                """
                UPDATE entries
                SET paper_received = ?, paper_received_at = ?,
                    paper_received_by = ?, paper_source = ?
                WHERE trans_no = ?
                """,
                (
                    1 if present else 0,
                    current if present else "",
                    user_name if present else "",
                    "manual-checkbox" if present else "",
                    trans_no,
                ),
            )
            audit(
                conn,
                "تحديث وجود الورق",
                "entry",
                trans_no,
                user_name,
                {"paper_received": bool(row["paper_received"])},
                {"paper_received": present},
                self.client_ip(),
            )
            conn.commit()
        self.send_json(
            {
                "ok": True,
                "trans_no": trans_no,
                "paper_received": present,
                "updated_by": user_name,
            }
        )

    def handle_attachment_from_url(self, trans_no: str) -> None:
        payload = read_json(self)
        url = str(payload.get("url") or "").strip()
        uploaded_by = self.request_user(
            explicit=str(payload.get("uploaded_by") or ""))
        if not uploaded_by:
            raise ValueError("اكتب اسم منفذ العملية الذي أضاف المرفق.")
        if not url.startswith(("http://", "https://")):
            raise ValueError(
                "اكتب رابط صورة أو PDF صحيح يبدأ بـ http أو https.")
        validate_remote_url(url)
        request = Request(url, headers={"User-Agent": "QoyodArchive/2.0"})
        with urlopen(request, timeout=25) as response:
            mime = response.headers.get_content_type() or "application/octet-stream"
            data = response.read(MAX_UPLOAD_BYTES +
                                 1) if MAX_UPLOAD_BYTES else response.read()
        if MAX_UPLOAD_BYTES and len(data) > MAX_UPLOAD_BYTES:
            raise ValueError(
                f"الملف من الرابط أكبر من الحد المسموح {MAX_UPLOAD_BYTES // 1024 // 1024}MB."
            )
        name = Path(
            urlparse(url).path).name or f"network-file-{int(time.time())}"
        with connect_db() as conn:
            attachment_id = store_attachment(
                conn,
                trans_no,
                name,
                data,
                mime,
                source="url",
                source_url=url,
                uploaded_by=uploaded_by,
                ip_address=self.client_ip(),
            )
            conn.commit()
        self.send_json(
            {"ok": True, "attachment_id": attachment_id, "filename": name})

    def handle_note(self, trans_no: str) -> None:
        payload = read_json(self)
        note = str(payload.get("note") or "").strip()
        author = self.request_user(explicit=str(payload.get("author") or ""))
        if not note:
            raise ValueError("اكتب الملاحظة.")
        if not author:
            raise ValueError("اكتب اسم صاحب الملاحظة.")
        with connect_db() as conn:
            if not conn.execute("SELECT 1 FROM entries WHERE trans_no = ?", (trans_no,)).fetchone():
                raise ValueError("رقم القيد غير موجود في كشف الأرشيف المعتمد.")
            conn.execute(
                "INSERT INTO notes (trans_no, author, role, note, created_at) VALUES (?, ?, ?, ?, ?)",
                (
                    trans_no,
                    author,
                    str(payload.get("role") or "").strip(),
                    note,
                    now_iso(),
                ),
            )
            audit(
                conn,
                "إضافة ملاحظة",
                "entry",
                trans_no,
                author,
                "",
                note,
                self.client_ip(),
            )
            conn.commit()
        self.send_json({"ok": True})

    def handle_file(self, attachment_id: int) -> None:
        with connect_db() as conn:
            row = conn.execute(
                "SELECT * FROM attachments WHERE id = ?", (attachment_id,)).fetchone()
        if not row:
            return self.send_error_json("المرفق غير موجود.", 404)
        path = attachment_path(row)
        if not path.exists():
            return self.send_error_json("ملف المرفق غير موجود على الهارد.", 404)
        self.record_activity(
            "فتح ملف مرفق",
            "attachment",
            attachment_id,
            {"trans_no": row["trans_no"], "file_name": row["original_name"]},
        )
        self.send_response(200)
        self.send_header(
            "Content-Type", row["mime"] or "application/octet-stream")
        self.send_header("Content-Length", str(path.stat().st_size))
        self.send_header("Content-Disposition",
                         f'inline; filename="{row["original_name"]}"')
        self.end_headers()
        with path.open("rb") as file:
            shutil.copyfileobj(file, self.wfile)

    def log_message(self, format: str, *args: object) -> None:
        LOGGER.info("%s - %s", self.address_string(), format % args)


def main() -> None:
    setup_logging()
    init_db()
    daily_backup = create_daily_backup()
    host = SETTINGS.host
    port = SETTINGS.port
    try:
        server = ThreadingHTTPServer((host, port), ArchiveHandler)
    except OSError as exc:
        LOGGER.error("تعذر تشغيل السيرفر على %s:%s: %s", host, port, exc)
        print(
            f"تعذر تشغيل السيرفر على المنفذ {port}. قد يكون مستخدمًا من برنامج آخر.")
        raise SystemExit(1) from exc
    print(f"Qoyod Archive is running on http://{host}:{port}")
    print(f"Data folder: {DATA_DIR}")
    print(f"Daily backup: {daily_backup}")
    LOGGER.info("Server started on http://%s:%s using %s", host, port, DB_PATH)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping...")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
